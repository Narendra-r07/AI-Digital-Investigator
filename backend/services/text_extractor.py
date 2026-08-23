from pathlib import Path


TEXT_EXTENSIONS = {
    ".txt",
    ".log",
    ".csv",
    ".md",
    ".json",
    ".xml",
    ".html",
    ".htm",
}


def extract_text(file_path: str) -> str:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"File not found: {file_path}"
        )

    extension = path.suffix.lower()

    # -------------------------------------------------
    # TEXT FILES
    # -------------------------------------------------

    if extension in TEXT_EXTENSIONS:
        return path.read_text(
            encoding="utf-8",
            errors="ignore",
        )

    # -------------------------------------------------
    # PDF
    # -------------------------------------------------

    if extension == ".pdf":
        try:
            import pypdf

            reader = pypdf.PdfReader(str(path))

            pages = []

            for page in reader.pages:
                pages.append(
                    page.extract_text() or ""
                )

            return "\n".join(pages)

        except ImportError:
            raise RuntimeError(
                "PDF extraction requires pypdf. "
                "Run: pip install pypdf"
            )

    # -------------------------------------------------
    # DOCX
    # -------------------------------------------------

    if extension == ".docx":
        try:
            from docx import Document

            document = Document(str(path))

            return "\n".join(
                paragraph.text
                for paragraph in document.paragraphs
                if paragraph.text.strip()
            )

        except ImportError:
            raise RuntimeError(
                "DOCX extraction requires python-docx. "
                "Run: pip install python-docx"
            )

    # -------------------------------------------------
    # XLSX
    # -------------------------------------------------

    if extension == ".xlsx":
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                filename=str(path),
                read_only=True,
                data_only=True,
            )

            output = []

            for sheet in workbook.worksheets:

                output.append(
                    f"--- Sheet: {sheet.title} ---"
                )

                for row in sheet.iter_rows(
                    values_only=True
                ):

                    values = [
                        str(value)
                        for value in row
                        if value is not None
                    ]

                    if values:
                        output.append(
                            " | ".join(values)
                        )

            return "\n".join(output)

        except ImportError:
            raise RuntimeError(
                "XLSX extraction requires openpyxl. "
                "Run: pip install openpyxl"
            )

    # -------------------------------------------------
    # Unsupported
    # -------------------------------------------------

    return ""